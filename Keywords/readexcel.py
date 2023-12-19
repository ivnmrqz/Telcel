import openpyxl,time
archivo=openpyxl.load_workbook("D:\\Proyectos GIT\\Telcel\\DATA\\numeros.xlsx")

def Numero_filas(hoja):
    ac=archivo[hoja]
    return ac.max_row

def Dato_columna(hoja,fila,col):
    ac=archivo[hoja]
    col=ac.cell(int(fila),int(col))
    return col.value

def escribe_mensaje(i,ms):
    ac=archivo["numeros"]
    msg=ac["B"+i].value=ms
    archivo.save("D:\\Proyectos GIT\\Telcel\\DATA\\numeros.xlsx")

def escribe_resultado(i,ms1,ms2,ms3,ms4,ms5):
    ac = archivo["numeros"]
    msg = ac["C" + i].value = ms1
    msg = ac["D" + i].value = ms2
    msg = ac["E" + i].value = ms3
    msg = ac["F" + i].value = ms4
    msg = ac["G" + i].value = ms5
    archivo.save("D:\\Proyectos GIT\\Telcel\\DATA\\numeros.xlsx")


def escribe_datos_renovacion(i,ms1,ms2,ms3,ms4,ms5,ms6,ms7,ms8,ms9,ms10,ms11,ms12,ms13,ms14,ms15,ms16):
    ac= archivo ["numeros"]
    msg = ac["H" + i].value = ms1
    msg = ac["I" + i].value = ms2
    msg = ac["J" + i].value = ms3
    msg = ac["K" + i].value = ms4
    msg = ac["L" + i].value = ms5
    msg = ac["M" + i].value = ms6
    msg = ac["N" + i].value = ms7
    msg = ac["O" + i].value = ms8
    msg = ac["P" + i].value = ms9
    msg = ac["Q" + i].value = ms10
    msg = ac["R" + i].value = ms11
    msg = ac["S" + i].value = ms12
    msg = ac["T" + i].value = ms13
    msg = ac["U" + i].value = ms14
    msg = ac["V" + i].value = ms15
    msg = ac["W" + i].value = ms16
    archivo.save("D:\\Proyectos GIT\\Telcel\\DATA\\numeros.xlsx")

